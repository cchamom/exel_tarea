import openpyxl

#paea poder agregar un gasto
def agregar_un_gasto(worksheet, fecha, descripcion, monto):
    ultima_fila = worksheet.max_row + 1

    worksheet.cell(row=ultima_fila, column=1, value=fecha)
    worksheet.cell(row=ultima_fila, column=2, value=descripcion)
    worksheet.cell(row=ultima_fila, column=3, value=monto)

#para calcular el resumen de los gastos
def calcular_el_resumen_exel(worksheet):
    numero_gastos = worksheet.max_row - 1  
    monto_total = sum(worksheet.cell(row=row, column=3).value for row in range(2, worksheet.max_row + 1))

    gasto_mas_caro = max(worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, values_only=True), key=lambda row: row[2])
    gasto_mas_barato = min(worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, values_only=True), key=lambda row: row[2])

    return numero_gastos, gasto_mas_caro, gasto_mas_barato, monto_total

#para poder crear el archivo de exel
def main():
    
    try:
        workbook = openpyxl.load_workbook("gastos.xlsx")
        worksheet = workbook["Gastos"]
    except FileNotFoundError:
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "Gastos"
        worksheet.append(["Fecha", "Descripción", "Monto"])

    while True:
        print("Precione la tecla 'n' si desea salir del programa")
        fecha = input("Ingrese la fecha del gasto: ")
        if fecha.lower() == 'n':
            break


        descrip = input("Ingrese la descripción del gasto: ")
        monto_gasto = float(input("Ingrese el monto del gasto: "))

        agregar_un_gasto(worksheet, fecha, descrip, monto_gasto)

    #para poder calcular el resumen
    numero_total_gastos, gasto_mas_caro, gasto_mas_barato, monto_total_gastos = calcular_el_resumen_exel(worksheet)

    print("\nResumen de los gastos:")
    print(f"Número total de gastos: {numero_total_gastos}")
    print(f"Gasto más caro: Fecha - {gasto_mas_caro[0]}, Descripción - {gasto_mas_caro[1]}, Monto - {gasto_mas_caro[2]}")
    print(f"Gasto más barato: Fecha - {gasto_mas_barato[0]}, Descripción - {gasto_mas_barato[1]}, Monto - {gasto_mas_barato[2]}")
    print(f"Monto total de gastos: {monto_total_gastos}")

    #guardar el archivo Excel
    workbook.save("gastos.xlsx")
    print("Infoorme de los gastos 'gastos.xlsx'")

if __name__ == "__main__":
    main()
